from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.database.models import (
    ApartmentUnit,
    ChessboardPriceEntry,
    ContractRegistryEntry,
    ResidentialComplex,
)


CHESS_HEADER_ALIASES: Dict[str, Iterable[str]] = {
    "block_name": ("блок", "block", "корпус"),
    "unit_type": ("тип", "type"),
    "status": ("статус", "status"),
    "rooms": ("колвокомнат", "количество комнат", "rooms"),
    "unit_number": ("номерпомещения", "номерпомещение", "номер", "квартира", "№кв", "№"),
    "area_sqm": ("площадьм2", "площадь", "area"),
    "floor": ("этаж", "floor"),
}


CONTRACT_HEADER_MAP: Dict[str, str] = {
    "договора": "contract_number",
    "недоговора": "contract_number",
    "датадоговора": "contract_date",
    "блок": "block_name",
    "этаж": "floor",
    "кв": "apartment_number",
    "квартиры": "apartment_number",
    "номерпомещения": "apartment_number",
    "колвоком": "rooms",
    "квадратураквартиры": "area_sqm",
    "общстоимостьдоговора": "total_price",
    "стоимость1квм": "price_per_sqm",
    "процент1взноса": "down_payment_percent",
    "сумма1взноса": "down_payment_amount",
    "ф/и/о": "buyer_full_name",
    "фио": "buyer_full_name",
    "серияпаспорта": "buyer_passport_series",
    "пинфл": "buyer_pinfl",
    "кемвыдан": "issued_by",
    "адреспрописки": "registration_address",
    "номертел": "phone_number",
    "отделпродаж": "sales_department",
}


DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d")


_CYR_TO_LAT = {
    "А": "A", "а": "a",
    "В": "V", "в": "v",
    "С": "S", "с": "s",
    "Е": "E", "е": "e",
    "К": "K", "к": "k",
    "М": "M", "м": "m",
    "Н": "N", "н": "n",
    "О": "O", "о": "o",
    "Р": "R", "р": "r",
    "Т": "T", "т": "t",
    "Х": "H", "х": "h",
    "У": "U", "у": "u",
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    cleaned = re.sub(r"[^a-zа-яё0-9%]+", "", str(value).strip().lower())
    return cleaned


def _normalize_block_name(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = "".join(_CYR_TO_LAT.get(ch, ch) for ch in s)
    s = s.lower()
    s = re.sub(r"[\s_–—-]+", "-", s)
    return s


def _normalize_unit_number(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    cleaned = str(value).strip().replace(",", ".")
    if cleaned == "":
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = (
        str(value)
        .replace(" ", "")
        .replace("\xa0", "")
        .replace(",", ".")
        .strip()
    )
    if cleaned == "":
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_date_value(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if text == "":
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _map_headers(headers: List[Any], aliases: Dict[str, Iterable[str]]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    normalized_headers = [_normalize_header(header) for header in headers]
    for field, patterns in aliases.items():
        for idx, header in enumerate(normalized_headers):
            if header and header in patterns:
                header_map[field] = idx
                break
    return header_map


def import_chess_from_excel(db: Session, complex_obj: ResidentialComplex, file_path: str) -> int:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    header_map = _map_headers(headers, {k: {_normalize_header(h) for h in v} for k, v in CHESS_HEADER_ALIASES.items()})

    required = {"block_name", "status", "unit_number", "floor"}
    if not required.issubset(header_map):
        missing = required - header_map.keys()
        raise ValueError(f"Отсутствуют обязательные колонки в шахматке: {', '.join(sorted(missing))}")

    db.query(ApartmentUnit).filter(ApartmentUnit.complex_id == complex_obj.id).delete(synchronize_session=False)

    new_units: List[ApartmentUnit] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        block_name = row[header_map["block_name"]]
        unit_number = row[header_map["unit_number"]]
        floor = row[header_map["floor"]]

        normalized_unit_number = _normalize_unit_number(unit_number)
        floor_int = _coerce_int(floor)

        if not normalized_unit_number or floor_int is None:
            continue

        payload = {str(headers[idx]): row[idx] for idx in range(len(headers)) if headers[idx] is not None}

        unit_type_value = None
        if "unit_type" in header_map:
            raw_type = row[header_map["unit_type"]]
            unit_type_value = str(raw_type).strip() if raw_type is not None else None

        status_raw = row[header_map["status"]]
        status_value = str(status_raw).strip() if status_raw is not None else ""

        rooms_value = _coerce_int(row[header_map["rooms"]]) if "rooms" in header_map else None
        area_value = _coerce_float(row[header_map["area_sqm"]]) if "area_sqm" in header_map else None

        new_units.append(
            ApartmentUnit(
                complex_id=complex_obj.id,
                block_name=str(block_name).strip() if block_name is not None else "",
                unit_type=unit_type_value,
                status=status_value,
                rooms=rooms_value,
                unit_number=normalized_unit_number,
                area_sqm=area_value,
                floor=floor_int,
                raw_payload=payload or None,
            )
        )

    if new_units:
        db.bulk_save_objects(new_units)

    return len(new_units)


def import_price_from_excel(db: Session, complex_obj: ResidentialComplex, file_path: str) -> int:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if not headers or len(headers) < 2:
        raise ValueError("Некорректный формат price_shaxamtka.xlsx: отсутствуют заголовки")

    category_headers = headers[1:]
    parsed_categories = [str(header) for header in category_headers]

    db.query(ChessboardPriceEntry).filter(ChessboardPriceEntry.complex_id == complex_obj.id).delete(synchronize_session=False)

    new_entries: List[ChessboardPriceEntry] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        floor_val = _coerce_int(row[0] if row else None)
        if floor_val is None:
            continue

        for order_index, category in enumerate(parsed_categories):
            price_val = _coerce_float(row[order_index + 1] if len(row) > order_index + 1 else None)
            if price_val is None:
                continue
            new_entries.append(
                ChessboardPriceEntry(
                    complex_id=complex_obj.id,
                    floor=floor_val,
                    category_key=str(category),
                    price_per_sqm=price_val,
                    order_index=order_index,
                )
            )

    if new_entries:
        db.bulk_save_objects(new_entries)

    return len(new_entries)


def _find_apartment_id(
        db: Session,
        complex_id: int,
        block_name: Any,
        floor: Optional[int],
        unit_number: Any,
) -> Optional[int]:
    if floor is None:
        return None
    normalized_number = _normalize_unit_number(unit_number)
    if not normalized_number:
        return None

    candidates = (
        db.query(ApartmentUnit)
        .filter(
            ApartmentUnit.complex_id == complex_id,
            ApartmentUnit.floor == floor,
            ApartmentUnit.unit_number == normalized_number,
        )
        .all()
    )

    target = next(
        (
            unit for unit in candidates
            if _normalize_block_name(unit.block_name) == _normalize_block_name(block_name)
        ),
        None,
    )
    return target.id if target else None


def import_contract_registry_from_excel(db: Session, complex_obj: ResidentialComplex, file_path: str) -> int:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    normalized_headers = [_normalize_header(header) for header in headers]

    db.query(ContractRegistryEntry).filter(ContractRegistryEntry.complex_id == complex_obj.id).delete(synchronize_session=False)

    new_entries: List[ContractRegistryEntry] = []
    seen_numbers: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue

        entry_data: Dict[str, Any] = {
            "complex_id": complex_obj.id,
            "extra_data": {},
        }

        for idx, header_key in enumerate(normalized_headers):
            value = row[idx] if idx < len(row) else None
            original_header = headers[idx] if idx < len(headers) else ""

            attr = CONTRACT_HEADER_MAP.get(header_key)
            if not attr:
                if original_header:
                    entry_data["extra_data"][original_header] = value
                continue

            if attr == "contract_date":
                parsed_date = _parse_date_value(value)
                entry_data[attr] = parsed_date or datetime.utcnow().date()
            elif attr in {"floor", "rooms"}:
                entry_data[attr] = _coerce_int(value)
            elif attr in {"total_price", "price_per_sqm", "down_payment_percent", "down_payment_amount", "area_sqm"}:
                entry_data[attr] = _coerce_float(value)
            elif attr == "apartment_number":
                entry_data[attr] = _normalize_unit_number(value)
            else:
                entry_data[attr] = str(value).strip() if value is not None else None

        if not entry_data.get("contract_number"):
            continue

        if not entry_data.get("buyer_full_name"):
            entry_data["buyer_full_name"] = ""

        if not entry_data.get("contract_date"):
            entry_data["contract_date"] = datetime.utcnow().date()

        contract_number = entry_data["contract_number"]
        if contract_number in seen_numbers:
            continue
        seen_numbers.add(contract_number)

        if "floor" in entry_data:
            apartment_id = _find_apartment_id(
                db,
                complex_obj.id,
                entry_data.get("block_name"),
                entry_data.get("floor"),
                entry_data.get("apartment_number"),
            )
            entry_data["apartment_id"] = apartment_id

        if not entry_data["extra_data"]:
            entry_data["extra_data"] = None

        new_entries.append(ContractRegistryEntry(**entry_data))

    if new_entries:
        db.bulk_save_objects(new_entries)

    return len(new_entries)
