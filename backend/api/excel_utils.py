from __future__ import annotations

import zipfile
from io import BytesIO
from typing import Union

from fastapi import APIRouter, HTTPException, Body
from num2words import num2words
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from pydantic import BaseModel, Field
import openpyxl
import os
import traceback  # Для логирования ошибок

from starlette.responses import StreamingResponse, FileResponse


# --- Вспомогательные функции ---
def col_letter_to_index(letter: str) -> int:
    """
    Преобразует буквенное обозначение столбца в индекс, используемый openpyxl (1-индексация).
    Например: 'A' -> 1, 'G' -> 7.
    """
    letter = letter.upper()
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index  # openpyxl использует 1-based index


# --- Конфигурация ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Словарь путей к файлам Excel для каждого ЖК
EXCEL_FILE_PATHS = {
    # !!! ЗАМЕНИТЕ НА ВАШИ РЕАЛЬНЫЕ ИМЕНА ЖК И ПУТИ !!!
    "ЖК_Бахор": os.path.join("static", "Жилые_Комплексы", "ЖК_Бахор", "jk_data.xlsx"),
    "Другой ЖК": os.path.join(BASE_DIR, "data", "shahmatka_drugoy.xlsx"),
}

# === НОВАЯ КОНФИГУРАЦИЯ СТОЛБЦОВ ===
COLUMN_MAPPING = {
    "blockName": 'A',  # Столбец 'A' для Блока
    "status": 'C',  # Столбец 'C' для Статуса (для обновления)
    "apartmentNumber": 'E',  # Столбец 'E' для Номера квартиры
    "floor": 'G'  # Столбец 'G' для Этажа
}
DATA_START_ROW = 2  # Строка, с которой начинаются данные (1-based)


# === КОНЕЦ НОВОЙ КОНФИГУРАЦИИ ===

# --- Модель запроса (остается без изменений) ---
class ApartmentStatusUpdate(BaseModel):
    jkName: str
    blockName: str
    floor: int
    apartmentNumber: Union[int, str]  # Номер квартиры может быть числом в JSON
    newStatus: str = Field(default="бронь")


class ContractData(BaseModel):
    jkName: str
    contractNumber: Union[str, None] = None
    contractDate: str
    block: str
    floor: int
    apartmentNumber: int
    rooms: int
    size: float
    paymentChoice: str
    totalPrice: str
    pricePerM2: str
    initialPayment: str
    fullName: str
    passportSeries: str
    pinfl: str
    issuedBy: str
    registrationAddress: str
    phone: str
    salesDepartment: str


# Функция преобразования числа в слова (на русском языке)
def number_to_words(number: str) -> str:
    # Убираем все нечисловые символы и преобразуем в int
    clean_number = ''.join(filter(str.isdigit, number))
    return num2words(int(clean_number), lang='ru') + " сум"


# --- Роутер (остается без изменений) ---
router = APIRouter(prefix="/api", tags=["Excel Operations"])  # Пример префикса


# --- Обновленная Логика обновления Excel ---
def find_row_and_update_status(file_path: str, update_data: ApartmentStatusUpdate):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        block_col_idx = col_letter_to_index(COLUMN_MAPPING["blockName"])
        floor_col_idx = col_letter_to_index(COLUMN_MAPPING["floor"])
        apt_num_col_idx = col_letter_to_index(COLUMN_MAPPING["apartmentNumber"])
        status_col_idx = col_letter_to_index(COLUMN_MAPPING["status"])

        print(f"Поиск в файле: {file_path}")
        print(f"Критерии: Блок='{update_data.blockName}', Этаж={update_data.floor}, Кв={update_data.apartmentNumber}")
        print(f"Колонки: Блок={block_col_idx}, Этаж={floor_col_idx}, Кв={apt_num_col_idx}, Статус={status_col_idx}")

        target_row_idx = -1
        for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
            cell_block = sheet.cell(row=row_idx, column=block_col_idx).value
            cell_floor = sheet.cell(row=row_idx, column=floor_col_idx).value
            cell_apt_num = sheet.cell(row=row_idx, column=apt_num_col_idx).value

            print(f"Строка {row_idx}: Блок='{cell_block}', Этаж='{cell_floor}', Кв='{cell_apt_num}'")

            try:
                current_block = str(cell_block).strip().lower() if cell_block else ""
                update_block = str(update_data.blockName).strip().lower()
                matches_block = current_block == update_block

                current_floor = int(cell_floor) if cell_floor is not None else None
                matches_floor = current_floor == update_data.floor

                current_apt_num = int(cell_apt_num) if cell_apt_num is not None else None
                matches_apt_num = current_apt_num == update_data.apartmentNumber

                if matches_block and matches_floor and matches_apt_num:
                    target_row_idx = row_idx
                    print(f"Найдена строка {row_idx}!")
                    break

            except (ValueError, TypeError) as e:
                print(f"Ошибка в строке {row_idx}: {e}")
                continue

        print(f"Итоговый target_row_idx: {target_row_idx}")
        if target_row_idx != -1:
            status_cell = sheet.cell(row=target_row_idx, column=status_col_idx)
            old_status = status_cell.value
            status_cell.value = update_data.newStatus
            workbook.save(file_path)
            print(f"Статус обновлен: строка {target_row_idx}, '{old_status}' -> '{update_data.newStatus}'")
            return True
        else:
            print(
                f"Квартира не найдена: Блок={update_data.blockName}, Этаж={update_data.floor}, Кв={update_data.apartmentNumber}")
            return False

    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Excel file not found: {file_path}")
    except Exception as e:
        print(f"Ошибка при обновлении Excel: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {e}")


@router.post("/excel/update-status")
async def update_excel_status_endpoint(update_data: ApartmentStatusUpdate = Body(...)):
    print(f"Запрос: {update_data.dict()}")
    file_path = EXCEL_FILE_PATHS.get(update_data.jkName)

    if not file_path:
        return {"status": "warning", "message": f"Excel file path config missing for '{update_data.jkName}'"}

    if not os.path.exists(file_path):
        return {"status": "warning", "message": f"Excel file not found for '{update_data.jkName}'"}

    success = find_row_and_update_status(file_path, update_data)
    if success:
        return {"status": "success", "message": "Apartment status updated successfully"}
    else:
        return {"status": "warning", "message": f"Apartment not found in Excel for '{update_data.jkName}'"}


BASE_STATIC_PATH = "static/Жилые_Комплексы"


@router.get("/last-contract-number")
async def get_last_contract_number(jkName: str):
    jk_dir = os.path.join(BASE_STATIC_PATH, jkName)
    REGISTRY_PATH = os.path.join(jk_dir, "contract_registry.xlsx")

    if not os.path.exists(REGISTRY_PATH):
        return {"lastContractNumber": None}  # Файл реестра не существует

    try:
        wb_registry = load_workbook(REGISTRY_PATH)
        ws_registry = wb_registry.active
        last_row = ws_registry.max_row

        if last_row <= 1:  # Только заголовок или пустой файл
            return {"lastContractNumber": None}

        last_contract_number = ws_registry[f"A{last_row}"].value  # Предполагается, что номер в колонке A
        return {"lastContractNumber": int(last_contract_number) + 1}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при чтении реестра: {str(e)}")


@router.post("/generate-contract")
async def generate_contract(data: ContractData):
    # Динамические пути к файлам в зависимости от jkName
    jk_dir = os.path.join(BASE_STATIC_PATH, data.jkName)
    TEMPLATE_PATH = os.path.join(jk_dir, "contract_template.xlsx")
    REGISTRY_PATH = os.path.join(jk_dir, "contract_registry.xlsx")

    # Проверка существования директории, если нет — создаем
    if not os.path.exists(jk_dir):
        os.makedirs(jk_dir)

    # 1. Генерация номера договора, если не указан
    if not data.contractNumber:
        if os.path.exists(REGISTRY_PATH):
            wb_registry = load_workbook(REGISTRY_PATH)
            ws_registry = wb_registry.active
            last_row = ws_registry.max_row
            last_contract_number = int(ws_registry[f"A{last_row}"].value.split('-')[1]) if last_row > 1 else 0
            next_contract_number = str(last_contract_number + 1).zfill(4)
        else:
            next_contract_number = "0001"
        data.contractNumber = f"Д-{next_contract_number}"

    # 2. Заполнение шаблона договора
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail=f"Шаблон договора не найден по пути: {TEMPLATE_PATH}")

    wb_contract = load_workbook(TEMPLATE_PATH)
    ws_contract = wb_contract.active

    # Замена placeholders
    replacements = {
        "{{Номер_Договора}}": data.contractNumber,
        "{{Дата}}": data.contractDate,
        "{{Ф_И_О}}": data.fullName,
        "{{Серия_Паспорта}}": data.passportSeries,
        "{{Кем_Выдан}}": data.issuedBy,
        "{{Прописка}}": data.registrationAddress,
        "{{Номер_Тел}}": data.phone,
        "{{ПИНФЛ}}": data.pinfl,
        "{{Блок}}": data.block,
        "{{Этаж}}": str(data.floor),
        "{{Номер_КВ}}": str(data.apartmentNumber),
        "{{Кол-во_Ком}}": str(data.rooms),
        "{{Квадратура_Квартиры}}": str(data.size),
        "{{Общ_Стоимость}}": data.totalPrice.replace(" ", ""),
        "{{Общ_Стоимость_Про}}": number_to_words(data.totalPrice),
        "{{Стоимость_1_м2}}": data.pricePerM2.replace(" ", ""),
        "{{Стоимость_1_м2_Про}}": number_to_words(data.pricePerM2),
        "{{Процент_1_Взноса}}": data.paymentChoice.replace("%", ""),
        "{{Сумма_1_Взноса}}": data.initialPayment.replace(" ", ""),
        "{{Сумма_1_Взноса_Про}}": number_to_words(data.initialPayment),
    }

    for row in ws_contract.rows:
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    cell.value = cell.value.replace(placeholder, value)

    # Сохранение заполненного договора
    contract_file = os.path.join(jk_dir, f"contract_{data.contractNumber}.xlsx")
    wb_contract.save(contract_file)

    # 3. Обновление реестра договоров
    if os.path.exists(REGISTRY_PATH):
        wb_registry = load_workbook(REGISTRY_PATH)
        ws_registry = wb_registry.active
    else:
        wb_registry = Workbook()
        ws_registry = wb_registry.active
        ws_registry.append([
            "№ Договора", "Дата Договора", "Блок", "Этаж", "№ КВ", "Кол-во ком",
            "Квадратура Квартиры", "Общ Стоимость Договора", "Стоимость 1 кв.м",
            "Процент 1 Взноса", "Сумма 1 Взноса", "Ф/И/О", "Серия Паспорта",
            "ПИНФЛ", "Кем выдан", "Адрес прописки", "Номер тел", "Отдел Продаж"
        ])

    ws_registry.append([
        data.contractNumber, data.contractDate, data.block, data.floor,
        data.apartmentNumber, data.rooms, data.size, data.totalPrice.replace(" ", ""),
        data.pricePerM2.replace(" ", ""), data.paymentChoice.replace("%", ""),
        data.initialPayment.replace(" ", ""), data.fullName, data.passportSeries,
        data.pinfl, data.issuedBy, data.registrationAddress, data.phone,
        data.salesDepartment
    ])

    wb_registry.save(REGISTRY_PATH)

    # 4. Создание ZIP-архива
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.write(contract_file, f"contract_{data.contractNumber}.xlsx")
        zip_file.write(REGISTRY_PATH, "contract_registry.xlsx")

    zip_buffer.seek(0)

    # 5. Удаление временного файла договора
    os.remove(contract_file)

    # 6. Отправка ZIP-архива клиенту
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=contract_and_registry_{data.contractNumber}.zip"
        }
    )


@router.get("/download-contract-registry")
async def download_contract_registry(jkName: str):
    jk_dir = os.path.join(BASE_STATIC_PATH, jkName)
    REGISTRY_PATH = os.path.join(jk_dir, "contract_registry.xlsx")

    if not os.path.exists(REGISTRY_PATH):
        raise HTTPException(status_code=404, detail="Реестр договоров не найден")

    return FileResponse(
        REGISTRY_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"contract_registry_{jkName}.xlsx"
    )


@router.get("/get-contract-registry")
async def get_contract_registry(jkName: str):
    jk_dir = os.path.join(BASE_STATIC_PATH, jkName)
    REGISTRY_PATH = os.path.join(jk_dir, "contract_registry.xlsx")

    if not os.path.exists(REGISTRY_PATH):
        raise HTTPException(status_code=404, detail="Реестр договоров не найден")

    try:
        wb_registry = load_workbook(REGISTRY_PATH)
        ws_registry = wb_registry.active
        data = []

        # Читаем заголовки из первой строки
        headers = [cell.value for cell in ws_registry[1]]

        # Читаем данные из остальных строк
        for row in ws_registry.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)

        return {"registry": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при чтении реестра: {str(e)}")
